import os
import traceback
import calendar
from datetime import datetime, timedelta
from io import BytesIO

from flask import Flask, render_template, request, redirect, send_file, url_for, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_connection

from database import get_connection

app = Flask(__name__)
app.secret_key = "kirakira_secret_key"
app.url_map.strict_slashes = False
DEBUG_TIENDA_LOGS = os.getenv("KIRAKIRA_DEBUG_TIENDA", "1") == "1"

def construir_hoja_kira(ws, nombre_tienda, ventas_por_fecha, anio, mes):
    fill_header = PatternFill(fill_type="solid", fgColor="5A9C9A")
    fill_card = PatternFill(fill_type="solid", fgColor="D9F0EF")

    font_header = Font(color="FFFFFF", bold=True)
    font_bold = Font(bold=True)
    font_title = Font(bold=True, size=14)

    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Anchos
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 18

    # Título
    ws["B1"] = f"VENTA MENSUAL {nombre_tienda}"
    ws["B1"].font = font_title

    # Encabezados tabla izquierda
    ws["B3"] = "Fecha"
    ws["C3"] = "Venta"

    ws["B3"].fill = fill_header
    ws["C3"].fill = fill_header
    ws["B3"].font = font_header
    ws["C3"].font = font_header
    ws["B3"].alignment = center
    ws["C3"].alignment = center
    ws["B3"].border = border
    ws["C3"].border = border

    # Todos los días del mes
    dias_mes = calendar.monthrange(anio, mes)[1]

    fila_inicio = 4
    fila = fila_inicio

    for dia in range(1, dias_mes + 1):
        fecha_str = f"{anio:04d}-{mes:02d}-{dia:02d}"
        venta = float(ventas_por_fecha.get(fecha_str, 0))

        ws[f"B{fila}"] = fecha_str
        ws[f"C{fila}"] = venta

        ws[f"B{fila}"].alignment = center
        ws[f"C{fila}"].alignment = center
        ws[f"B{fila}"].border = border
        ws[f"C{fila}"].border = border
        ws[f"C{fila}"].number_format = '"$"#,##0.00'

        fila += 1

    fila_fin_ventas = fila - 1

    # Bloque resumen
    ws["E3"] = "Resumen"
    ws["F3"] = "Monto"
    ws["E3"].fill = fill_header
    ws["F3"].fill = fill_header
    ws["E3"].font = font_header
    ws["F3"].font = font_header
    ws["E3"].alignment = center
    ws["F3"].alignment = center
    ws["E3"].border = border
    ws["F3"].border = border

    resumen_labels = [
        "Ventas del mes",
        "Renta",
        "Mantenimiento",
        "Luz",
        "Sueldos",
        "Gastos variables",
        "Total gastos fijos",
        "Gasto total",
        "Ganancia total"
    ]

    for i, label in enumerate(resumen_labels, start=4):
        ws[f"E{i}"] = label
        ws[f"F{i}"].fill = fill_card
        ws[f"E{i}"].fill = fill_card
        ws[f"E{i}"].font = font_bold
        ws[f"E{i}"].alignment = left
        ws[f"F{i}"].alignment = center
        ws[f"E{i}"].border = border
        ws[f"F{i}"].border = border
        ws[f"F{i}"].number_format = '"$"#,##0.00'

    # Fórmulas
    ws["F4"] = f"=SUM(C{fila_inicio}:C{fila_fin_ventas})"  # Ventas del mes
    ws["F5"] = 0  # Renta
    ws["F6"] = 0  # Mantenimiento
    ws["F7"] = 0  # Luz
    ws["F8"] = 0  # Sueldos
    ws["F9"] = 0  # Gastos variables manuales
    ws["F10"] = "=SUM(F5:F8)"   # Total gastos fijos
    ws["F11"] = "=F10+F9"       # Gasto total
    ws["F12"] = "=F4-F11"       # Ganancia total

def _debug_log(message):
    if DEBUG_TIENDA_LOGS:
        app.logger.info(message)


def _safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _tienda_nombre(tienda_id):
    return f"Tienda {tienda_id}"


def _carrito_key(tienda_id):
    return f"carrito_t{tienda_id}"


def _session_tienda_id():
    return _safe_int(session.get("tienda_id"))


def _normalize_role(rol):
    return (rol or "").strip().lower()


def _is_admin_role(rol):
    return _normalize_role(rol) in {"administrador", "admin"}


def _is_store_staff_role(rol):
    return _normalize_role(rol) in {"empleado", "vendedor", "vendedora"}


def _employee_dashboard_endpoint(tienda_id):
    if tienda_id == 1:
        return "tienda1_dashboard"
    if tienda_id == 2:
        return "tienda2_dashboard"
    return None


def _store_back_url(tienda_id):
    if _is_admin_role(session.get("rol")):
        return url_for("admin_balance")

    endpoint = _employee_dashboard_endpoint(tienda_id)
    if endpoint:
        return url_for(endpoint)

    return url_for("home")


def _require_store_access(tienda_id, route_name):
    if "usuario" not in session:
        return redirect(url_for("home"))

    rol = session.get("rol")
    tienda_session = _session_tienda_id()

    _debug_log(
        f"[CTX] route={route_name} usuario={session.get('usuario')} "
        f"rol={rol} tienda_session={tienda_session} tienda_forzada={tienda_id}"
    )

    if _is_admin_role(rol):
        return None

    if _is_store_staff_role(rol) and tienda_session == tienda_id:
        return None

    if _is_store_staff_role(rol) and tienda_session in (1, 2):
        endpoint = _employee_dashboard_endpoint(tienda_session)
        if endpoint:
            return redirect(url_for(endpoint))

    return "No autorizado para esta tienda", 403


def _log_query(route_name, query_name, tienda_id, params):
    _debug_log(
        f"[SQL] route={route_name} query={query_name} tienda_id={tienda_id} params={params}"
    )


def _render_caja_tienda(tienda_id, template_name, route_name):
    access = _require_store_access(tienda_id, route_name)
    if access is not None:
        return access

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, codigo, nombre, precio, stock FROM productos ORDER BY nombre ASC")
    productos = cursor.fetchall()
    conn.close()

    carrito = session.get(_carrito_key(tienda_id), [])
    total = sum(item["subtotal"] for item in carrito)

    return render_template(
        template_name,
        tienda_id=tienda_id,
        tienda_nombre=_tienda_nombre(tienda_id),
        back_url=_store_back_url(tienda_id),
        productos=productos,
        carrito=carrito,
        total=total,
    )


def _agregar_carrito_tienda(tienda_id, route_name, redirect_endpoint):
    access = _require_store_access(tienda_id, route_name)
    if access is not None:
        return access

    conn = get_connection()
    cursor = conn.cursor()

    producto_id = request.form["producto_id"]
    cantidad = _safe_int(request.form.get("cantidad"))

    if cantidad is None:
        conn.close()
        return "Cantidad invalida"

    cursor.execute(
        """
        SELECT id, nombre, precio, stock
        FROM productos
        WHERE id = ?
    """,
        (producto_id,),
    )
    producto = cursor.fetchone()
    conn.close()

    if not producto:
        return "Producto no encontrado"

    if cantidad <= 0:
        return "La cantidad debe ser mayor a 0"

    if cantidad > producto[3]:
        return f"No hay suficiente stock. Disponible: {producto[3]}"

    key = _carrito_key(tienda_id)
    carrito = session.get(key, [])
    subtotal = producto[2] * cantidad

    carrito.append(
        {
            "producto_id": producto[0],
            "nombre": producto[1],
            "precio": producto[2],
            "cantidad": cantidad,
            "subtotal": subtotal,
        }
    )

    session[key] = carrito
    return redirect(url_for(redirect_endpoint))


def _eliminar_carrito_tienda(tienda_id, route_name, redirect_endpoint, index):
    access = _require_store_access(tienda_id, route_name)
    if access is not None:
        return access

    key = _carrito_key(tienda_id)
    carrito = session.get(key, [])

    if 0 <= index < len(carrito):
        carrito.pop(index)

    session[key] = carrito
    return redirect(url_for(redirect_endpoint))


def _finalizar_venta_tienda(tienda_id, route_name, redirect_endpoint):
    access = _require_store_access(tienda_id, route_name)
    if access is not None:
        return access

    key = _carrito_key(tienda_id)
    carrito = session.get(key, [])

    if not carrito:
        return "El carrito esta vacio"

    metodo_pago = request.form.get("metodo_pago", "").strip()
    if not metodo_pago:
        return "Metodo de pago invalido"

    conn = get_connection()
    cursor = conn.cursor()

    total = sum(item["subtotal"] for item in carrito)
    _log_query(route_name, "venta_insert", tienda_id, (total, metodo_pago, tienda_id))

    cursor.execute(
        """
        INSERT INTO ventas (total, metodo_pago, tienda_id)
        VALUES (?, ?, ?)
    """,
        (total, metodo_pago, tienda_id),
    )
    venta_id = cursor.lastrowid

    for item in carrito:
        _log_query(
            route_name,
            "venta_detalle_insert",
            tienda_id,
            (venta_id, item["producto_id"], item["cantidad"], item["precio"], item["subtotal"]),
        )
        cursor.execute(
            """
            INSERT INTO venta_detalle (venta_id, producto_id, cantidad, precio_unitario, subtotal)
            VALUES (?, ?, ?, ?, ?)
        """,
            (
                venta_id,
                item["producto_id"],
                item["cantidad"],
                item["precio"],
                item["subtotal"],
            ),
        )

        _log_query(
            route_name,
            "producto_stock_update",
            tienda_id,
            (item["cantidad"], item["producto_id"]),
        )
        cursor.execute(
            """
            UPDATE productos
            SET stock = stock - ?
            WHERE id = ?
        """,
            (item["cantidad"], item["producto_id"]),
        )

    conn.commit()
    conn.close()

    session[key] = []
    return redirect(url_for(redirect_endpoint))


def _ventas_dia_tienda(tienda_id, template_name, route_name):
    access = _require_store_access(tienda_id, route_name)
    if access is not None:
        return access

    conn = get_connection()
    cursor = conn.cursor()

    _log_query(route_name, "ventas_hoy", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT id, total, metodo_pago, fecha
        FROM ventas
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
        ORDER BY fecha DESC
    """,
        (tienda_id,),
    )
    ventas_hoy = cursor.fetchall()

    _log_query(route_name, "total_vendido", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT COALESCE(SUM(total), 0)
        FROM ventas
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
    """,
        (tienda_id,),
    )
    total_vendido = cursor.fetchone()[0]

    _log_query(route_name, "total_gastos", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT COALESCE(SUM(monto), 0)
        FROM gastos_caja
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
    """,
        (tienda_id,),
    )
    total_gastos = cursor.fetchone()[0]

    balance_neto = total_vendido - total_gastos
    conn.close()

    return render_template(
        template_name,
        tienda_id=tienda_id,
        tienda_nombre=_tienda_nombre(tienda_id),
        back_url=_store_back_url(tienda_id),
        ventas_hoy=ventas_hoy,
        total_vendido=total_vendido,
        total_gastos=total_gastos,
        balance_neto=balance_neto,
    )


def _gastos_tienda(tienda_id, template_name, route_name):
    access = _require_store_access(tienda_id, route_name)
    if access is not None:
        return access

    conn = get_connection()
    cursor = conn.cursor()
    mensaje = ""

    if request.method == "POST":
        concepto = request.form.get("concepto", "").strip()
        monto_texto = request.form.get("monto", "").strip()

        if not concepto:
            mensaje = "Debes capturar un concepto."
        else:
            try:
                monto = float(monto_texto)
            except ValueError:
                mensaje = "Monto invalido."
            else:
                if monto <= 0:
                    mensaje = "El monto debe ser mayor a 0."
                else:
                    _log_query(route_name, "gasto_insert", tienda_id, (concepto, monto, tienda_id))
                    cursor.execute(
                        """
                        INSERT INTO gastos_caja (concepto, monto, tienda_id)
                        VALUES (?, ?, ?)
                    """,
                        (concepto, monto, tienda_id),
                    )
                    conn.commit()
                    mensaje = "Gasto guardado correctamente"

    _log_query(route_name, "gastos_hoy", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT id, concepto, monto, fecha
        FROM gastos_caja
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
        ORDER BY id DESC
    """,
        (tienda_id,),
    )
    gastos = cursor.fetchall()

    _log_query(route_name, "total_gastos", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT COALESCE(SUM(monto), 0)
        FROM gastos_caja
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
    """,
        (tienda_id,),
    )
    total_gastos = cursor.fetchone()[0]

    conn.close()

    return render_template(
        template_name,
        tienda_id=tienda_id,
        tienda_nombre=_tienda_nombre(tienda_id),
        back_url=_store_back_url(tienda_id),
        gastos=gastos,
        total_gastos=total_gastos,
        mensaje=mensaje,
    )


def _cierre_tienda(tienda_id, template_name, route_name):
    access = _require_store_access(tienda_id, route_name)
    if access is not None:
        return access

    conn = get_connection()
    cursor = conn.cursor()
    mensaje = ""
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")

    if request.method == "POST":
        monto_texto = request.form.get("monto", "").strip()

        try:
            monto = float(monto_texto)
        except ValueError:
            mensaje = "Monto invalido."
        else:
            if monto < 0:
                mensaje = "El monto no puede ser negativo."
            else:
                _log_query(route_name, "fondo_buscar", tienda_id, (fecha_hoy, tienda_id))
                cursor.execute(
                    """
                    SELECT id
                    FROM fondo_caja
                    WHERE fecha = ? AND tienda_id = ?
                    ORDER BY id DESC
                    LIMIT 1
                """,
                    (fecha_hoy, tienda_id),
                )
                existente = cursor.fetchone()

                if existente:
                    _log_query(route_name, "fondo_update", tienda_id, (monto, existente[0]))
                    cursor.execute(
                        """
                        UPDATE fondo_caja
                        SET monto = ?
                        WHERE id = ?
                    """,
                        (monto, existente[0]),
                    )
                else:
                    _log_query(route_name, "fondo_insert", tienda_id, (fecha_hoy, monto, tienda_id))
                    cursor.execute(
                        """
                        INSERT INTO fondo_caja (fecha, monto, tienda_id)
                        VALUES (?, ?, ?)
                    """,
                        (fecha_hoy, monto, tienda_id),
                    )

                conn.commit()
                mensaje = "Fondo guardado correctamente."

    _log_query(route_name, "ventas_hoy", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT id, total, metodo_pago, fecha
        FROM ventas
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
        ORDER BY fecha DESC
    """,
        (tienda_id,),
    )
    ventas_hoy = cursor.fetchall()

    _log_query(route_name, "total_vendido", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT COALESCE(SUM(total), 0)
        FROM ventas
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
    """,
        (tienda_id,),
    )
    total_vendido = cursor.fetchone()[0]

    _log_query(route_name, "total_gastos", tienda_id, (tienda_id,))
    cursor.execute(
        """
        SELECT COALESCE(SUM(monto), 0)
        FROM gastos_caja
        WHERE DATE(fecha, 'localtime') = DATE('now', 'localtime')
          AND tienda_id = ?
    """,
        (tienda_id,),
    )
    total_gastos = cursor.fetchone()[0]

    balance_neto = total_vendido - total_gastos

    _log_query(route_name, "monto_fondo", tienda_id, (fecha_hoy, tienda_id))
    cursor.execute(
        """
        SELECT COALESCE(SUM(monto), 0)
        FROM fondo_caja
        WHERE fecha = ? AND tienda_id = ?
    """,
        (fecha_hoy, tienda_id),
    )
    monto_fondo = cursor.fetchone()[0]

    conn.close()

    return render_template(
        template_name,
        tienda_id=tienda_id,
        tienda_nombre=_tienda_nombre(tienda_id),
        back_url=_store_back_url(tienda_id),
        ventas_hoy=ventas_hoy,
        total_vendido=total_vendido,
        total_gastos=total_gastos,
        balance_neto=balance_neto,
        monto_fondo=monto_fondo,
        mensaje=mensaje,
    )


def _legacy_redirect(modulo):
    if "usuario" not in session:
        return redirect(url_for("home"))

    rol = session.get("rol")
    if _is_admin_role(rol):
        return redirect(url_for("admin_balance"))

    if not _is_store_staff_role(rol):
        return "Rol sin acceso de tienda", 403

    tienda_id = _session_tienda_id()
    if tienda_id not in (1, 2):
        return "Empleado sin tienda asignada"

    endpoint = None
    if modulo == "caja":
        endpoint = f"tienda{tienda_id}_caja"
    elif modulo == "gastos":
        endpoint = f"tienda{tienda_id}_gastos"
    elif modulo == "ventas":
        endpoint = f"tienda{tienda_id}_ventas_dia"
    elif modulo == "cierre":
        endpoint = f"tienda{tienda_id}_cierre"

    if endpoint is None:
        return redirect(url_for("home"))

    return redirect(url_for(endpoint))


@app.route("/", methods=["GET"], endpoint="home")
def home():
    return render_template("login.html")

@app.route("/admin/exportar-excel-30")
def exportar_excel_30():
    conn = get_connection()
    cursor = conn.cursor()

    hoy = datetime.now().date()
    anio = hoy.year
    mes = hoy.month
    inicio_mes = f"{anio:04d}-{mes:02d}-01"
    fin_mes = f"{anio:04d}-{mes:02d}-{calendar.monthrange(anio, mes)[1]:02d}"

    # KIRA 1
    cursor.execute("""
        SELECT DATE(fecha), COALESCE(SUM(total), 0)
        FROM ventas
        WHERE tienda_id = 1 AND DATE(fecha) BETWEEN ? AND ?
        GROUP BY DATE(fecha)
        ORDER BY DATE(fecha)
    """, (inicio_mes, fin_mes))
    kira1_raw = cursor.fetchall()
    kira1_dict = {str(fecha): float(total) for fecha, total in kira1_raw}

    # KIRA 2
    cursor.execute("""
        SELECT DATE(fecha), COALESCE(SUM(total), 0)
        FROM ventas
        WHERE tienda_id = 2 AND DATE(fecha) BETWEEN ? AND ?
        GROUP BY DATE(fecha)
        ORDER BY DATE(fecha)
    """, (inicio_mes, fin_mes))
    kira2_raw = cursor.fetchall()
    kira2_dict = {str(fecha): float(total) for fecha, total in kira2_raw}

    conn.close()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "KIRA 1"
    construir_hoja_kira(ws1, "KIRA 1", kira1_dict, anio, mes)

    ws2 = wb.create_sheet(title="KIRA 2")
    construir_hoja_kira(ws2, "KIRA 2", kira2_dict, anio, mes)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"reporte_mensual_kira_{anio}_{mes:02d}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/health", methods=["GET"], endpoint="health")
def health():
    try:
        conn = get_connection()
        conn.close()
        return "Conexion exitosa usando SQLite"
    except Exception as e:
        print(traceback.format_exc())
        return f"Error de conexion: {repr(e)}"


@app.route("/consultar-productos", endpoint="consultar_productos")
def consultar_productos():
    if "usuario" not in session:
        return redirect(url_for("home"))

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT id, nombre, codigo, precio, stock, categoria
        FROM productos
        WHERE COALESCE(activo, 1) = 1
        ORDER BY nombre ASC
    """
    )
    productos = cursor.fetchall()
    conn.close()

    return render_template("consultar_productos.html", productos=productos)


@app.route("/admin-dashboard", endpoint="admin_dashboard")
def admin_dashboard():
    return redirect(url_for("admin_balance"))

@app.route("/tienda1/inventario", methods=["GET", "POST"], endpoint="tienda1_inventario")
def tienda1_inventario():
    conn = get_connection()
    cursor = conn.cursor()
    mensaje = None
    error = None

    if request.method == "POST":
        accion = request.form.get("accion")

        # =========================
        # AGREGAR PRODUCTO NUEVO
        # =========================
        if accion == "nuevo_producto":
            codigo = request.form.get("codigo", "").strip()
            nombre = request.form.get("nombre", "").strip()
            precio = request.form.get("precio", "").strip()
            stock_inicial = request.form.get("stock_inicial", "").strip()

            if not codigo or not nombre or not precio or not stock_inicial:
                error = "Completa cÃƒÂ³digo, nombre, precio y stock inicial."
            else:
                try:
                    precio = float(precio)
                    stock_inicial = int(stock_inicial)

                    if precio < 0 or stock_inicial < 0:
                        error = "Precio y stock no pueden ser negativos."
                    else:
                        cursor.execute(
                            "SELECT id FROM productos WHERE codigo = ?",
                            (codigo,)
                        )
                        existente = cursor.fetchone()

                        if existente:
                            error = "Ya existe un producto con ese cÃƒÂ³digo."
                        else:
                            cursor.execute(
                                "INSERT INTO productos (codigo, nombre, precio, stock) VALUES (?, ?, ?, ?)",
                                (codigo, nombre, precio, stock_inicial)
                            )
                            conn.commit()
                            mensaje = "Producto agregado correctamente."
                except ValueError:
                    error = "Precio o stock invÃƒÂ¡lidos."

        # =========================
        # MOVIMIENTO DE INVENTARIO
        # =========================
        elif accion == "movimiento":
            producto_id = request.form.get("producto_id")
            tipo = request.form.get("tipo")
            cantidad_txt = request.form.get("cantidad", "").strip()

            if not producto_id or not cantidad_txt:
                error = "Selecciona producto y cantidad."
            else:
                try:
                    cantidad = int(cantidad_txt)

                    if cantidad <= 0:
                        error = "La cantidad debe ser mayor a 0."
                    else:
                        cursor.execute(
                            "SELECT stock, nombre FROM productos WHERE id = ?",
                            (producto_id,)
                        )
                        producto = cursor.fetchone()

                        if not producto:
                            error = "Producto no encontrado."
                        else:
                            stock_actual = producto[0]
                            nombre_producto = producto[1]

                            if tipo == "entrada":
                                nuevo_stock = stock_actual + cantidad
                            elif tipo == "salida":
                                if cantidad > stock_actual:
                                    error = f"No hay suficiente inventario de {nombre_producto}."
                                else:
                                    nuevo_stock = stock_actual - cantidad
                            else:
                                error = "Tipo de movimiento invÃƒÂ¡lido."

                            if not error:
                                cursor.execute(
                                    "UPDATE productos SET stock = ? WHERE id = ?",
                                    (nuevo_stock, producto_id)
                                )
                                conn.commit()
                                mensaje = "Movimiento guardado correctamente."
                except ValueError:
                    error = "Cantidad invÃƒÂ¡lida."

        elif accion == "eliminar_producto":
            producto_id = request.form.get("producto_id_eliminar")
            confirmacion = request.form.get("confirmacion_eliminar", "").strip().upper()

            if not producto_id:
                error = "Selecciona un producto para eliminar."
            elif confirmacion != "ELIMINAR":
                error = "Escribe ELIMINAR para confirmar la eliminacion."
            else:
                cursor.execute(
                    "SELECT nombre FROM productos WHERE id = ?",
                    (producto_id,),
                )
                producto = cursor.fetchone()

                if not producto:
                    error = "Producto no encontrado."
                else:
                    nombre_producto = producto[0]
                    try:
                        cursor.execute(
                            "DELETE FROM productos WHERE id = ?",
                            (producto_id,),
                        )
                        conn.commit()
                        mensaje = f"Producto {nombre_producto} eliminado definitivamente."

                        key = _carrito_key(1)
                        carrito = session.get(key, [])
                        session[key] = [
                            item
                            for item in carrito
                            if str(item.get("producto_id")) != str(producto_id)
                        ]
                    except Exception:
                        conn.rollback()
                        error = "No se pudo eliminar el producto. Verifica si tiene ventas relacionadas."

    cursor.execute("SELECT id, codigo, nombre, precio, stock FROM productos ORDER BY nombre ASC")
    productos = cursor.fetchall()

    conn.close()
    return render_template(
        "inventario.html",
        productos=productos,
        mensaje=mensaje,
        error=error,
        admin_balance_url=url_for("admin_balance"),
        titulo_tienda="Tienda 1"
    )


@app.route("/tienda2/inventario", methods=["GET", "POST"], endpoint="tienda2_inventario")
def tienda2_inventario():
    conn = get_connection()
    cursor = conn.cursor()
    mensaje = None
    error = None

    if request.method == "POST":
        accion = request.form.get("accion")

        # =========================
        # AGREGAR PRODUCTO NUEVO
        # =========================
        if accion == "nuevo_producto":
            codigo = request.form.get("codigo", "").strip()
            nombre = request.form.get("nombre", "").strip()
            precio = request.form.get("precio", "").strip()
            stock_inicial = request.form.get("stock_inicial", "").strip()

            if not codigo or not nombre or not precio or not stock_inicial:
                error = "Completa cÃƒÂ³digo, nombre, precio y stock inicial."
            else:
                try:
                    precio = float(precio)
                    stock_inicial = int(stock_inicial)

                    if precio < 0 or stock_inicial < 0:
                        error = "Precio y stock no pueden ser negativos."
                    else:
                        cursor.execute(
                            "SELECT id FROM productos WHERE codigo = ?",
                            (codigo,)
                        )
                        existente = cursor.fetchone()

                        if existente:
                            error = "Ya existe un producto con ese cÃƒÂ³digo."
                        else:
                            cursor.execute(
                                "INSERT INTO productos (codigo, nombre, precio, stock) VALUES (?, ?, ?, ?)",
                                (codigo, nombre, precio, stock_inicial)
                            )
                            conn.commit()
                            mensaje = "Producto agregado correctamente."
                except ValueError:
                    error = "Precio o stock invÃƒÂ¡lidos."

        # =========================
        # MOVIMIENTO DE INVENTARIO
        # =========================
        elif accion == "movimiento":
            producto_id = request.form.get("producto_id")
            tipo = request.form.get("tipo")
            cantidad_txt = request.form.get("cantidad", "").strip()

            if not producto_id or not cantidad_txt:
                error = "Selecciona producto y cantidad."
            else:
                try:
                    cantidad = int(cantidad_txt)

                    if cantidad <= 0:
                        error = "La cantidad debe ser mayor a 0."
                    else:
                        cursor.execute(
                            "SELECT stock, nombre FROM productos WHERE id = ?",
                            (producto_id,)
                        )
                        producto = cursor.fetchone()

                        if not producto:
                            error = "Producto no encontrado."
                        else:
                            stock_actual = producto[0]
                            nombre_producto = producto[1]

                            if tipo == "entrada":
                                nuevo_stock = stock_actual + cantidad
                            elif tipo == "salida":
                                if cantidad > stock_actual:
                                    error = f"No hay suficiente inventario de {nombre_producto}."
                                else:
                                    nuevo_stock = stock_actual - cantidad
                            else:
                                error = "Tipo de movimiento invÃƒÂ¡lido."

                            if not error:
                                cursor.execute(
                                    "UPDATE productos SET stock = ? WHERE id = ?",
                                    (nuevo_stock, producto_id)
                                )
                                conn.commit()
                                mensaje = "Movimiento guardado correctamente."
                except ValueError:
                    error = "Cantidad invÃƒÂ¡lida."

        elif accion == "eliminar_producto":
            producto_id = request.form.get("producto_id_eliminar")
            confirmacion = request.form.get("confirmacion_eliminar", "").strip().upper()

            if not producto_id:
                error = "Selecciona un producto para eliminar."
            elif confirmacion != "ELIMINAR":
                error = "Escribe ELIMINAR para confirmar la eliminacion."
            else:
                cursor.execute(
                    "SELECT nombre FROM productos WHERE id = ?",
                    (producto_id,),
                )
                producto = cursor.fetchone()

                if not producto:
                    error = "Producto no encontrado."
                else:
                    nombre_producto = producto[0]
                    try:
                        cursor.execute(
                            "DELETE FROM productos WHERE id = ?",
                            (producto_id,),
                        )
                        conn.commit()
                        mensaje = f"Producto {nombre_producto} eliminado definitivamente."

                        key = _carrito_key(2)
                        carrito = session.get(key, [])
                        session[key] = [
                            item
                            for item in carrito
                            if str(item.get("producto_id")) != str(producto_id)
                        ]
                    except Exception:
                        conn.rollback()
                        error = "No se pudo eliminar el producto. Verifica si tiene ventas relacionadas."

    cursor.execute("SELECT id, codigo, nombre, precio, stock FROM productos ORDER BY nombre ASC")
    productos = cursor.fetchall()
    
    conn.close()
    return render_template(
        "inventario.html",
        productos=productos,
        mensaje=mensaje,
        error=error,
        admin_balance_url=url_for("admin_balance"),
        titulo_tienda="Tienda 2"
    )


@app.route("/empleado-dashboard", endpoint="empleado_dashboard")
def empleado_dashboard():
    if "usuario" not in session:
        return redirect(url_for("home"))

    rol = session.get("rol")
    if _is_admin_role(rol):
        return redirect(url_for("admin_balance"))

    if not _is_store_staff_role(rol):
        return "Rol no autorizado"

    tienda_id = _session_tienda_id()
    endpoint = _employee_dashboard_endpoint(tienda_id)
    if endpoint:
        return redirect(url_for(endpoint))

    return "Empleado sin tienda asignada"

from datetime import datetime, timedelta

from datetime import datetime, timedelta
from flask import render_template, redirect, url_for, session

@app.route("/admin/balance", endpoint="admin_balance")
def admin_balance():
    if "usuario" not in session:
        return redirect(url_for("home"))

    conn = get_connection()
    cursor = conn.cursor()

    hoy = datetime.now().date()
    hace_15 = hoy - timedelta(days=14)
    hace_30 = hoy - timedelta(days=29)

    # =========================
    # VENTAS HOY
    # =========================
    cursor.execute("""
        SELECT COALESCE(SUM(total), 0)
        FROM ventas
        WHERE DATE(fecha) = ?
    """, (hoy,))
    ventas_hoy = cursor.fetchone()[0]

    # =========================
    # GASTOS HOY
    # =========================
    cursor.execute("""
        SELECT COALESCE(SUM(monto), 0)
        FROM gastos_caja
        WHERE DATE(fecha) = ?
    """, (hoy,))
    gastos_hoy = cursor.fetchone()[0]

    neto_hoy = ventas_hoy - gastos_hoy

    # =========================
    # ÃƒÅ¡LTIMOS 15 DÃƒÂAS
    # =========================
    cursor.execute("""
        SELECT COALESCE(SUM(total), 0)
        FROM ventas
        WHERE DATE(fecha) BETWEEN ? AND ?
    """, (hace_15, hoy))
    ventas_15 = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COALESCE(SUM(monto), 0)
        FROM gastos_caja
        WHERE DATE(fecha) BETWEEN ? AND ?
    """, (hace_15, hoy))
    gastos_15 = cursor.fetchone()[0]

    neto_15 = ventas_15 - gastos_15

    # =========================
    # ÃƒÅ¡LTIMOS 30 DÃƒÂAS
    # =========================
    cursor.execute("""
        SELECT COALESCE(SUM(total), 0)
        FROM ventas
        WHERE DATE(fecha) BETWEEN ? AND ?
    """, (hace_30, hoy))
    ventas_30 = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COALESCE(SUM(monto), 0)
        FROM gastos_caja
        WHERE DATE(fecha) BETWEEN ? AND ?
    """, (hace_30, hoy))
    gastos_30 = cursor.fetchone()[0]

    neto_30 = ventas_30 - gastos_30

    fechas = [(hoy - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(15)]
    fechas.reverse()

    cursor.execute(
        """
        SELECT
            DATE(fecha) AS fecha,
            tienda_id,
            COALESCE(SUM(total), 0) AS ventas
        FROM ventas
        WHERE DATE(fecha) BETWEEN ? AND ?
          AND tienda_id IN (1, 2)
        GROUP BY DATE(fecha), tienda_id
        ORDER BY DATE(fecha), tienda_id
    """,
        (hace_15, hoy),
    )
    ventas_rows = cursor.fetchall()

    cursor.execute(
        """
        SELECT
            DATE(fecha) AS fecha,
            tienda_id,
            COALESCE(SUM(monto), 0) AS gastos
        FROM gastos_caja
        WHERE DATE(fecha) BETWEEN ? AND ?
          AND tienda_id IN (1, 2)
        GROUP BY DATE(fecha), tienda_id
        ORDER BY DATE(fecha), tienda_id
    """,
        (hace_15, hoy),
    )
    gastos_rows = cursor.fetchall()

    ventas_por_fecha_tienda = {}
    for fecha, tienda_id, ventas in ventas_rows:
        if tienda_id in (1, 2):
            ventas_por_fecha_tienda[(fecha, tienda_id)] = ventas or 0

    gastos_por_fecha_tienda = {}
    for fecha, tienda_id, gastos in gastos_rows:
        if tienda_id in (1, 2):
            gastos_por_fecha_tienda[(fecha, tienda_id)] = gastos or 0

    def construir_reporte_tienda(tienda_id):
        reporte = []
        for fecha in fechas:
            ventas = ventas_por_fecha_tienda.get((fecha, tienda_id), 0) or 0
            gastos = gastos_por_fecha_tienda.get((fecha, tienda_id), 0) or 0
            ganancia = ventas - gastos
            reporte.append(
                {
                    "fecha": fecha,
                    "ventas": ventas,
                    "gastos": gastos,
                    "ganancia": ganancia,
                }
            )
        return reporte

    reporte_kira1 = construir_reporte_tienda(1)
    reporte_kira2 = construir_reporte_tienda(2)

    conn.close()

    return render_template(
        "admin_balance.html",
        ventas_hoy=ventas_hoy,
        gastos_hoy=gastos_hoy,
        neto_hoy=neto_hoy,
        ventas_15=ventas_15,
        gastos_15=gastos_15,
        neto_15=neto_15,
        ventas_30=ventas_30,
        gastos_30=gastos_30,
        neto_30=neto_30,
        reporte_kira1=reporte_kira1,
        reporte_kira2=reporte_kira2
    )


@app.route("/tienda1", endpoint="tienda1_root")
def tienda1_root():
    return redirect(url_for("tienda1_dashboard"))


@app.route("/tienda2", endpoint="tienda2_root")
def tienda2_root():
    return redirect(url_for("tienda2_dashboard"))


@app.route("/tienda1/dashboard", endpoint="tienda1_dashboard")
def tienda1_dashboard():
    access = _require_store_access(1, "tienda1_dashboard")
    if access is not None:
        return access
    return render_template("tienda1/dashboard.html", tienda_nombre="Tienda 1")


@app.route("/tienda2/dashboard", endpoint="tienda2_dashboard")
def tienda2_dashboard():
    access = _require_store_access(2, "tienda2_dashboard")
    if access is not None:
        return access
    return render_template("tienda2/dashboard.html", tienda_nombre="Tienda 2")


@app.route("/tienda1/caja", endpoint="tienda1_caja")
def tienda1_caja():
    return _render_caja_tienda(1, "tienda1/caja.html", "tienda1_caja")


@app.route("/tienda2/caja", endpoint="tienda2_caja")
def tienda2_caja():
    return _render_caja_tienda(2, "tienda2/caja.html", "tienda2_caja")


@app.route("/tienda1/caja/agregar", methods=["POST"], endpoint="tienda1_caja_agregar")
def tienda1_caja_agregar():
    return _agregar_carrito_tienda(1, "tienda1_caja_agregar", "tienda1_caja")


@app.route("/tienda2/caja/agregar", methods=["POST"], endpoint="tienda2_caja_agregar")
def tienda2_caja_agregar():
    return _agregar_carrito_tienda(2, "tienda2_caja_agregar", "tienda2_caja")


@app.route("/tienda1/caja/eliminar/<int:index>", endpoint="tienda1_caja_eliminar")
def tienda1_caja_eliminar(index):
    return _eliminar_carrito_tienda(1, "tienda1_caja_eliminar", "tienda1_caja", index)


@app.route("/tienda2/caja/eliminar/<int:index>", endpoint="tienda2_caja_eliminar")
def tienda2_caja_eliminar(index):
    return _eliminar_carrito_tienda(2, "tienda2_caja_eliminar", "tienda2_caja", index)


@app.route("/tienda1/caja/finalizar", methods=["POST"], endpoint="tienda1_caja_finalizar")
def tienda1_caja_finalizar():
    return _finalizar_venta_tienda(1, "tienda1_caja_finalizar", "tienda1_caja")


@app.route("/tienda2/caja/finalizar", methods=["POST"], endpoint="tienda2_caja_finalizar")
def tienda2_caja_finalizar():
    return _finalizar_venta_tienda(2, "tienda2_caja_finalizar", "tienda2_caja")


@app.route("/tienda1/ventas-dia", endpoint="tienda1_ventas_dia")
def tienda1_ventas_dia():
    return _ventas_dia_tienda(1, "tienda1/ventas_dia.html", "tienda1_ventas_dia")


@app.route("/tienda2/ventas-dia", endpoint="tienda2_ventas_dia")
def tienda2_ventas_dia():
    return _ventas_dia_tienda(2, "tienda2/ventas_dia.html", "tienda2_ventas_dia")


@app.route("/tienda1/gastos", methods=["GET", "POST"], endpoint="tienda1_gastos")
def tienda1_gastos():
    return _gastos_tienda(1, "tienda1/gastos.html", "tienda1_gastos")


@app.route("/tienda2/gastos", methods=["GET", "POST"], endpoint="tienda2_gastos")
def tienda2_gastos():
    return _gastos_tienda(2, "tienda2/gastos.html", "tienda2_gastos")


@app.route("/tienda1/cierre", methods=["GET", "POST"], endpoint="tienda1_cierre")
def tienda1_cierre():
    return _cierre_tienda(1, "tienda1/cierre.html", "tienda1_cierre")


@app.route("/tienda2/cierre", methods=["GET", "POST"], endpoint="tienda2_cierre")
def tienda2_cierre():
    return _cierre_tienda(2, "tienda2/cierre.html", "tienda2_cierre")


@app.route("/caja", endpoint="caja")
def caja_legacy():
    return _legacy_redirect("caja")


@app.route("/ventas_del_dia", endpoint="ventas_del_dia")
@app.route("/ventas-dia", endpoint="ventas_dia")
def ventas_legacy():
    return _legacy_redirect("ventas")


@app.route("/gastos-caja", methods=["GET", "POST"], endpoint="gastos_caja")
def gastos_legacy():
    return _legacy_redirect("gastos")


@app.route("/cierre_caja", methods=["GET", "POST"], endpoint="cierre_caja")
def cierre_legacy():
    return _legacy_redirect("cierre")

@app.route("/agregar-al-carrito", methods=["POST"], endpoint="agregar_al_carrito")
def agregar_al_carrito_legacy():
    tienda_id = _session_tienda_id()
    if tienda_id == 1:
        return tienda1_caja_agregar()
    if tienda_id == 2:
        return tienda2_caja_agregar()
    return redirect(url_for("home"))


@app.route("/eliminar-del-carrito/<int:index>", endpoint="eliminar_del_carrito")
def eliminar_del_carrito_legacy(index):
    tienda_id = _session_tienda_id()
    if tienda_id == 1:
        return tienda1_caja_eliminar(index)
    if tienda_id == 2:
        return tienda2_caja_eliminar(index)
    return redirect(url_for("home"))


@app.route("/finalizar-venta", methods=["POST"], endpoint="finalizar_venta")
def finalizar_venta_legacy():
    tienda_id = _session_tienda_id()
    if tienda_id == 1:
        return tienda1_caja_finalizar()
    if tienda_id == 2:
        return tienda2_caja_finalizar()
    return redirect(url_for("home"))


@app.route("/login", methods=["POST"], endpoint="login")
def login():
    usuario = request.form["usuario"]
    password = request.form["password"]

    try:
        conn = get_connection()
        cur = conn.cursor()

        _debug_log(f"[AUTH] intento_login usuario={usuario}")
        cur.execute(
            """
            SELECT id, nombre, usuario, rol, tienda_id
            FROM usuarios
            WHERE usuario = ? AND password = ?
        """,
            (usuario, password),
        )
        user = cur.fetchone()

        cur.close()
        conn.close()

        if not user:
            return "Usuario o contrasena incorrectos"

        rol = user[3]
        tienda_id = _safe_int(user[4])

        session.clear()
        session["usuario"] = user[2]
        session["rol"] = rol
        session["tienda_id"] = tienda_id
        session[_carrito_key(1)] = []
        session[_carrito_key(2)] = []

        _debug_log(
            f"[AUTH] login_ok usuario={session.get('usuario')} rol={rol} tienda_session={session.get('tienda_id')}"
        )

        if _is_admin_role(rol):
            return redirect(url_for("admin_balance"))

        if _is_store_staff_role(rol):
            endpoint = _employee_dashboard_endpoint(tienda_id)
            if endpoint:
                return redirect(url_for(endpoint))
            return "Empleado sin tienda asignada"

        return f"Rol no reconocido: {rol}"

    except Exception as e:
        print(traceback.format_exc())
        return f"Error en login: {repr(e)}"


if __name__ == "__main__":
    port = int(os.getenv("APP_PORT", "5055"))
    app.run(debug=True, port=port, use_reloader=False)
